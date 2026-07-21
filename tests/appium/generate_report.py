"""
generate_report.py — Appium Mobile E2E report generator.
Produces Excel (.xlsx), HTML (dark-themed), and Markdown summary.

Usage:
    python tests/appium/generate_report.py \
        --json  test-results.json \
        --base-url https://user.github.io/medibook/
"""

import argparse
import json
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from jinja2 import Environment, BaseLoader
    HAS_JINJA2 = True
except ImportError:
    HAS_JINJA2 = False

ROOT        = Path(__file__).parent
RESULTS_DIR = ROOT / "Test Results"
EXCEL_DIR   = RESULTS_DIR / "Excel"
HTML_DIR    = RESULTS_DIR / "HTML"
SUMMARY_DIR = RESULTS_DIR / "Summary"

for d in [EXCEL_DIR, HTML_DIR, SUMMARY_DIR]:
    d.mkdir(parents=True, exist_ok=True)


# ─────────────────────────────────────────────────────────────
# Parse pytest JSON
# ─────────────────────────────────────────────────────────────
def parse_json(json_path: str) -> dict:
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    tests = []
    for t in data.get("tests", []):
        outcome = t.get("outcome", "unknown")
        call    = t.get("call", {})
        tests.append({
            "id":       t.get("nodeid", ""),
            "name":     t.get("nodeid", "").split("::")[-1],
            "module":   t.get("nodeid", "").split("::")[0].replace("tests/appium/", ""),
            "outcome":  outcome,
            "duration": round(call.get("duration", 0), 2),
            "error":    str(call.get("longrepr", ""))[:300] if outcome == "failed" else "",
        })
    s = data.get("summary", {})
    return {
        "tests":        tests,
        "total":        s.get("total",   len(tests)),
        "passed":       s.get("passed",  sum(1 for t in tests if t["outcome"] == "passed")),
        "failed":       s.get("failed",  sum(1 for t in tests if t["outcome"] == "failed")),
        "skipped":      s.get("skipped", sum(1 for t in tests if t["outcome"] == "skipped")),
        "duration":     round(data.get("duration", 0), 2),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ─────────────────────────────────────────────────────────────
# Excel report
# ─────────────────────────────────────────────────────────────
def generate_excel(report: dict, base_url: str, run_number: str = "N/A") -> str:
    if not HAS_OPENPYXL:
        print("[WARN] openpyxl not installed — skipping Excel.")
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    PURPLE = "7C3AED"; GREEN = "10B981"; RED = "EF4444"
    AMBER  = "F59E0B"; DARK  = "1E1B4B"; WHITE = "FFFFFF"; GRAY = "F1F5F9"

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "MediBook — Android Appium Mobile E2E Test Report"
    c.font  = Font(bold=True, size=16, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    meta = [
        ("Test Type",       "Appium Mobile E2E (Chrome on Android)"),
        ("Deployment URL",  base_url),
        ("Build Number",    run_number),
        ("Generated At",    report["generated_at"]),
        ("Total Duration",  f"{report['duration']} s"),
    ]
    for i, (lbl, val) in enumerate(meta, 2):
        ws.cell(i, 1, lbl).font = Font(bold=True, color=DARK)
        ws.cell(i, 2, val)

    # KPI row
    ws.merge_cells("A8:G8")
    ws["A8"].value = "EXECUTION SUMMARY"
    ws["A8"].font  = Font(bold=True, size=12, color=WHITE)
    ws["A8"].fill  = PatternFill("solid", fgColor=PURPLE)
    ws["A8"].alignment = Alignment(horizontal="center")

    pct  = round(report["passed"] / report["total"] * 100, 1) if report["total"] else 0
    kpis = [
        ("Total",   report["total"],   "4F46E5"),
        ("Passed",  report["passed"],  GREEN),
        ("Failed",  report["failed"],  RED),
        ("Skipped", report["skipped"], AMBER),
        ("Pass %",  f"{pct}%",         PURPLE),
    ]
    for col, (lbl, val, clr) in enumerate(kpis, 1):
        ws.cell(9,  col, lbl).font = Font(bold=True, color=WHITE)
        ws.cell(9,  col, lbl).fill = PatternFill("solid", fgColor=clr)
        ws.cell(9,  col, lbl).alignment = Alignment(horizontal="center")
        ws.cell(10, col, val).font = Font(bold=True, size=14, color=clr)
        ws.cell(10, col, val).alignment = Alignment(horizontal="center")
    ws.row_dimensions[10].height = 30

    # Details sheet
    ws2 = wb.create_sheet("Test Details")
    hdr = ["#", "Test ID", "Test Name", "Module", "Outcome", "Duration (s)", "Error"]
    for col, h in enumerate(hdr, 1):
        c = ws2.cell(1, col, h)
        c.font  = Font(bold=True, color=WHITE)
        c.fill  = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center")

    OC = {"passed": GREEN, "failed": RED, "skipped": AMBER}
    for row, t in enumerate(report["tests"], 2):
        clr = OC.get(t["outcome"], "888888")
        row_data = [row - 1, t["id"], t["name"], t["module"],
                    t["outcome"].upper(), t["duration"], t["error"]]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row, col, val)
            if col == 5:
                cell.font = Font(bold=True, color=clr)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GRAY)

    for sheet in [ws, ws2]:
        for col in sheet.columns:
            w = max((len(str(c.value or "")) for c in col), default=0)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 60)

    out = str(EXCEL_DIR / "Automation_Test_Report.xlsx")
    wb.save(out)
    print(f"[OK] Excel -> {out}")
    return out


# ─────────────────────────────────────────────────────────────
# HTML report
# ─────────────────────────────────────────────────────────────
HTML_TMPL = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1.0"/>
  <title>MediBook Appium E2E Report — Build #{{ run_number }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0f0f1a;color:#e2e8f0;padding:2rem}
    h1{font-size:2rem;background:linear-gradient(135deg,#7c3aed,#2563eb);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:.25rem}
    .subtitle{color:#64748b;font-size:.9rem;margin-bottom:2rem}
    .meta-bar{display:flex;gap:2rem;flex-wrap:wrap;background:#1e1b4b;border-radius:10px;padding:1rem 1.5rem;margin-bottom:2rem;font-size:.85rem;border:1px solid #312e81}
    .meta-bar span{color:#64748b}.meta-bar strong{color:#a5b4fc}
    .kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:1rem;margin-bottom:2rem}
    .kpi{background:#1e1b4b;border-radius:12px;padding:1.25rem;text-align:center;border:1px solid #312e81}
    .kpi .val{font-size:2.25rem;font-weight:800}
    .kpi .lbl{font-size:.7rem;text-transform:uppercase;letter-spacing:.1em;color:#94a3b8;margin-top:.25rem}
    .kpi.pass .val{color:#10b981}.kpi.fail .val{color:#ef4444}.kpi.skip .val{color:#f59e0b}.kpi.pct .val{color:#7c3aed}.kpi.total .val{color:#60a5fa}
    .progress-bar{background:#1e293b;border-radius:9999px;height:8px;overflow:hidden;margin-bottom:2rem}
    .progress-fill{height:100%;background:linear-gradient(90deg,#10b981,#34d399);border-radius:9999px;transition:width .5s}
    table{width:100%;border-collapse:collapse;background:#13131f;border-radius:12px;overflow:hidden;margin-bottom:2rem}
    th{background:#1e1b4b;padding:.75rem 1rem;text-align:left;font-size:.75rem;text-transform:uppercase;letter-spacing:.05em;color:#94a3b8}
    td{padding:.75rem 1rem;border-bottom:1px solid #1e293b;font-size:.825rem}
    tr:last-child td{border-bottom:none}tr:hover td{background:#1a1a2e}
    .badge{display:inline-block;padding:.2rem .6rem;border-radius:9999px;font-size:.75rem;font-weight:700}
    .badge.passed{background:#064e3b;color:#10b981}.badge.failed{background:#450a0a;color:#ef4444}.badge.skipped{background:#451a03;color:#f59e0b}
    .error{color:#ef4444;font-size:.75rem;white-space:pre-wrap;max-height:60px;overflow:auto;background:#1a0a0a;border-radius:6px;padding:.5rem;margin-top:.25rem}
    .url-box{background:#1e1b4b;border-radius:8px;padding:.75rem 1rem;border-left:4px solid #7c3aed;font-family:monospace;color:#7c3aed;font-size:.85rem;margin-bottom:2rem;word-break:break-all}
    .report-link{display:inline-block;background:linear-gradient(135deg,#7c3aed,#4f46e5);color:#fff;padding:.6rem 1.25rem;border-radius:8px;text-decoration:none;font-weight:600;font-size:.875rem;margin-bottom:2rem}
    footer{text-align:center;color:#475569;font-size:.8rem;margin-top:2rem;padding-top:1rem;border-top:1px solid #1e293b}
  </style>
</head>
<body>
  <h1>MediBook Appium E2E Report</h1>
  <p class="subtitle">Android Chrome Mobile Testing · Build #{{ run_number }} · {{ generated_at }}</p>

  <div class="url-box">🔗 {{ base_url }}</div>
  <a class="report-link" href="../../../reports/index.html">← All Reports</a>

  <div class="meta-bar">
    <div><span>Platform </span><strong>Android Chrome</strong></div>
    <div><span>Driver </span><strong>Appium + UiAutomator2</strong></div>
    <div><span>Duration </span><strong>{{ duration }}s</strong></div>
    <div><span>Build </span><strong>#{{ run_number }}</strong></div>
  </div>

  <div class="kpi-grid">
    <div class="kpi total"><div class="val">{{ total }}</div><div class="lbl">Total</div></div>
    <div class="kpi pass"> <div class="val">{{ passed }}</div><div class="lbl">Passed</div></div>
    <div class="kpi fail"> <div class="val">{{ failed }}</div><div class="lbl">Failed</div></div>
    <div class="kpi skip"> <div class="val">{{ skipped }}</div><div class="lbl">Skipped</div></div>
    <div class="kpi pct">  <div class="val">{{ pct }}%</div><div class="lbl">Pass Rate</div></div>
  </div>

  <div class="progress-bar">
    <div class="progress-fill" style="width:{{ pct }}%"></div>
  </div>

  <table>
    <thead><tr><th>#</th><th>Test Name</th><th>Module</th><th>Result</th><th>Duration</th><th>Details</th></tr></thead>
    <tbody>
    {% for t in tests %}
    <tr>
      <td>{{ loop.index }}</td>
      <td>{{ t.name }}</td>
      <td style="color:#64748b;font-size:.75rem">{{ t.module }}</td>
      <td><span class="badge {{ t.outcome }}">{{ t.outcome.upper() }}</span></td>
      <td>{{ t.duration }}s</td>
      <td>{% if t.error %}<div class="error">{{ t.error }}</div>{% endif %}</td>
    </tr>
    {% endfor %}
    </tbody>
  </table>

  <footer>MediBook · Appium Mobile E2E · Phase 7 CI/CD · Build #{{ run_number }}</footer>
</body>
</html>"""


def generate_html(report: dict, base_url: str, run_number: str = "N/A") -> str:
    if not HAS_JINJA2:
        out = str(HTML_DIR / "execution-report.html")
        with open(out, "w", encoding="utf-8") as f:
            f.write(f"<html><body><h1>MediBook Appium Report</h1>"
                    f"<p>Total: {report['total']} Passed: {report['passed']} Failed: {report['failed']}</p></body></html>")
        return out

    pct = round(report["passed"] / report["total"] * 100, 1) if report["total"] else 0
    env  = Environment(loader=BaseLoader())
    tmpl = env.from_string(HTML_TMPL)
    html = tmpl.render(**report, base_url=base_url, pct=pct, run_number=run_number)
    out  = str(HTML_DIR / "execution-report.html")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] HTML -> {out}")
    return out


# ─────────────────────────────────────────────────────────────
# Markdown summary
# ─────────────────────────────────────────────────────────────
def generate_summary(report: dict, base_url: str, run_number: str = "N/A") -> str:
    pct    = round(report["passed"] / report["total"] * 100, 1) if report["total"] else 0
    failed = [t for t in report["tests"] if t["outcome"] == "failed"]
    report_url = f"{base_url.rstrip('/')}/reports/latest/execution-report.html"

    lines = [
        "# Android Appium Test Summary\n",
        f"**Build Number:** {run_number}  ",
        f"**Execution Date:** {report['generated_at']}  ",
        f"**Platform:** Android Chrome (Appium + UiAutomator2)  \n",
        "## Results\n",
        "| Metric | Value |", "|--------|-------|",
        f"| Total Tests | {report['total']} |",
        f"| Passed | ✅ {report['passed']} |",
        f"| Failed | ❌ {report['failed']} |",
        f"| Skipped | ⏭️ {report['skipped']} |",
        f"| Pass Rate | **{pct}%** |",
        f"| Duration | {report['duration']}s |\n",
        f"## Report URL\n",
        f"🔗 [{report_url}]({report_url})\n",
    ]
    if failed:
        lines += ["\n## Failed Tests\n"]
        for t in failed:
            msg = t["error"][:200].replace("\n", " ") if t["error"] else "assertion error"
            lines.append(f"- **{t['name']}** — {msg}")
    else:
        lines.append("\n✅ All tests passed!")

    out = str(SUMMARY_DIR / "summary.md")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"[OK] Summary -> {out}")
    return out


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--json",       default="test-results.json")
    parser.add_argument("--base-url",   default="http://localhost:5173")
    parser.add_argument("--run-number", default="local")
    args = parser.parse_args()

    if os.path.exists(args.json):
        report = parse_json(args.json)
    else:
        print(f"[WARN] {args.json} not found — empty report.")
        report = {"tests": [], "total": 0, "passed": 0, "failed": 0,
                  "skipped": 0, "duration": 0,
                  "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

    generate_excel(report, args.base_url, args.run_number)
    generate_html(report,  args.base_url, args.run_number)
    generate_summary(report, args.base_url, args.run_number)
    print(f"\n[DONE] All reports in: {RESULTS_DIR}")


if __name__ == "__main__":
    main()
