"""
generate_report.py — Creates Excel + HTML reports from pytest's JSON output.

Usage (called automatically by CI workflow):
    python tests/generate_report.py --json test-results.json --base-url <url>
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

# ── Try to import optional deps ──────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from jinja2 import Environment, BaseLoader
    HAS_JINJA2 = True
except ImportError:
    HAS_JINJA2 = False

# ── Output directories ────────────────────────────────────────
ROOT         = Path(__file__).parent
RESULTS_DIR  = ROOT / "Test Results"
EXCEL_DIR    = RESULTS_DIR / "Excel"
HTML_DIR     = RESULTS_DIR / "HTML"
SUMMARY_DIR  = RESULTS_DIR / "Summary"

for d in [EXCEL_DIR, HTML_DIR, SUMMARY_DIR]:
    d.mkdir(parents=True, exist_ok=True)


# ─────────────────────────────────────────────────────────────
# Parse pytest JSON report
# ─────────────────────────────────────────────────────────────
def parse_json_report(json_path: str) -> dict:
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    tests = []
    for t in data.get("tests", []):
        outcome = t.get("outcome", "unknown")
        call    = t.get("call", {})
        tests.append({
            "id":       t.get("nodeid", ""),
            "name":     t.get("nodeid", "").split("::")[-1],
            "module":   t.get("nodeid", "").split("::")[0],
            "outcome":  outcome,
            "duration": round(t.get("call", {}).get("duration", 0), 2),
            "error":    call.get("longrepr", "") if outcome == "failed" else "",
        })

    summary = data.get("summary", {})
    return {
        "tests":    tests,
        "total":    summary.get("total",   len(tests)),
        "passed":   summary.get("passed",  sum(1 for t in tests if t["outcome"] == "passed")),
        "failed":   summary.get("failed",  sum(1 for t in tests if t["outcome"] == "failed")),
        "skipped":  summary.get("skipped", sum(1 for t in tests if t["outcome"] == "skipped")),
        "duration": round(data.get("duration", 0), 2),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ─────────────────────────────────────────────────────────────
# Excel report
# ─────────────────────────────────────────────────────────────
def generate_excel(report: dict, base_url: str) -> str:
    if not HAS_OPENPYXL:
        print("[WARN] openpyxl not installed — skipping Excel report.")
        return ""

    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    PURPLE = "7C3AED"
    GREEN  = "10B981"
    RED    = "EF4444"
    AMBER  = "F59E0B"
    DARK   = "1E1B4B"
    WHITE  = "FFFFFF"
    LGRAY  = "F1F5F9"

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "MediBook — Live GitHub Pages E2E Test Report"
    title_cell.font      = Font(bold=True, size=16, color=WHITE)
    title_cell.fill      = PatternFill("solid", fgColor=DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Metadata rows
    meta = [
        ("Deployment URL", base_url),
        ("Generated At",   report["generated_at"]),
        ("Total Duration", f"{report['duration']} s"),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        ws.cell(i, 1, label).font  = Font(bold=True, color=DARK)
        ws.cell(i, 2, value).font  = Font(color=DARK)

    # KPI row
    ws.merge_cells("A6:F6")
    ws["A6"].value = "TEST EXECUTION SUMMARY"
    ws["A6"].font  = Font(bold=True, size=12, color=WHITE)
    ws["A6"].fill  = PatternFill("solid", fgColor=PURPLE)
    ws["A6"].alignment = Alignment(horizontal="center")

    pct  = round(report["passed"] / report["total"] * 100, 1) if report["total"] else 0
    kpis = [
        ("Total",   report["total"],   "4F46E5"),
        ("Passed",  report["passed"],  "10B981"),
        ("Failed",  report["failed"],  "EF4444"),
        ("Skipped", report["skipped"], "F59E0B"),
        ("Pass %",  f"{pct}%",         "7C3AED"),
    ]
    for col, (label, value, colour) in enumerate(kpis, start=1):
        ws.cell(7, col, label).font  = Font(bold=True, color=WHITE)
        ws.cell(7, col, label).fill  = PatternFill("solid", fgColor=colour)
        ws.cell(7, col, label).alignment = Alignment(horizontal="center")
        ws.cell(8, col, value).font  = Font(bold=True, size=14, color=colour)
        ws.cell(8, col, value).alignment = Alignment(horizontal="center")
        ws.row_dimensions[8].height = 28

    # ── Test Details sheet ────────────────────────────────────
    ws2 = wb.create_sheet("Test Details")
    headers = ["#", "Test ID", "Test Name", "Module", "Outcome", "Duration (s)", "Error"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(1, col, h)
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="center")

    OUTCOME_COLOURS = {"passed": GREEN, "failed": RED, "skipped": AMBER}
    for row, t in enumerate(report["tests"], start=2):
        colour = OUTCOME_COLOURS.get(t["outcome"], "888888")
        data   = [row - 1, t["id"], t["name"], t["module"],
                  t["outcome"].upper(), t["duration"], t["error"][:200] if t["error"] else ""]
        for col, val in enumerate(data, 1):
            cell = ws2.cell(row, col, val)
            if col == 5:  # Outcome column
                cell.font = Font(bold=True, color=colour)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LGRAY)

    # Auto-width
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    out_path = str(EXCEL_DIR / "Automation_Test_Report.xlsx")
    wb.save(out_path)
    print(f"[OK] Excel report -> {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────
# HTML report
# ─────────────────────────────────────────────────────────────
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>MediBook E2E Test Report</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0f0f1a;color:#e2e8f0;padding:2rem}
    h1{font-size:2rem;background:linear-gradient(135deg,#7c3aed,#2563eb);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:.5rem}
    .meta{color:#64748b;font-size:.875rem;margin-bottom:2rem}
    .kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:1rem;margin-bottom:2rem}
    .kpi{background:#1e1b4b;border-radius:12px;padding:1.5rem;text-align:center}
    .kpi .val{font-size:2.5rem;font-weight:800}
    .kpi .lbl{font-size:.75rem;text-transform:uppercase;letter-spacing:.1em;color:#94a3b8;margin-top:.25rem}
    .kpi.pass .val{color:#10b981}
    .kpi.fail .val{color:#ef4444}
    .kpi.skip .val{color:#f59e0b}
    .kpi.pct .val{color:#7c3aed}
    .kpi.total .val{color:#60a5fa}
    table{width:100%;border-collapse:collapse;background:#13131f;border-radius:12px;overflow:hidden}
    th{background:#1e1b4b;padding:.75rem 1rem;text-align:left;font-size:.75rem;text-transform:uppercase;letter-spacing:.05em;color:#94a3b8}
    td{padding:.75rem 1rem;border-bottom:1px solid #1e293b;font-size:.875rem}
    tr:last-child td{border-bottom:none}
    tr:hover td{background:#1a1a2e}
    .badge{display:inline-block;padding:.2rem .6rem;border-radius:9999px;font-size:.75rem;font-weight:700}
    .badge.passed{background:#064e3b;color:#10b981}
    .badge.failed{background:#450a0a;color:#ef4444}
    .badge.skipped{background:#451a03;color:#f59e0b}
    .error{color:#ef4444;font-size:.8rem;white-space:pre-wrap;max-height:80px;overflow:auto}
    .url-bar{background:#1e1b4b;border-radius:8px;padding:.75rem 1rem;margin-bottom:1.5rem;font-family:monospace;color:#7c3aed;border-left:4px solid #7c3aed}
    footer{margin-top:2rem;text-align:center;color:#64748b;font-size:.8rem}
  </style>
</head>
<body>
  <h1>MediBook — Live E2E Test Report</h1>
  <p class="meta">Generated: {{ generated_at }} | Duration: {{ duration }}s</p>
  <div class="url-bar">🔗 {{ base_url }}</div>

  <div class="kpi-grid">
    <div class="kpi total"><div class="val">{{ total }}</div><div class="lbl">Total</div></div>
    <div class="kpi pass"> <div class="val">{{ passed }}</div><div class="lbl">Passed</div></div>
    <div class="kpi fail"> <div class="val">{{ failed }}</div><div class="lbl">Failed</div></div>
    <div class="kpi skip"> <div class="val">{{ skipped }}</div><div class="lbl">Skipped</div></div>
    <div class="kpi pct">  <div class="val">{{ pct }}%</div><div class="lbl">Pass Rate</div></div>
  </div>

  <table>
    <thead>
      <tr>
        <th>#</th><th>Test Name</th><th>Module</th>
        <th>Outcome</th><th>Duration</th><th>Error</th>
      </tr>
    </thead>
    <tbody>
      {% for t in tests %}
      <tr>
        <td>{{ loop.index }}</td>
        <td>{{ t.name }}</td>
        <td style="color:#64748b">{{ t.module }}</td>
        <td><span class="badge {{ t.outcome }}">{{ t.outcome.upper() }}</span></td>
        <td>{{ t.duration }}s</td>
        <td>{% if t.error %}<div class="error">{{ t.error[:300] }}</div>{% endif %}</td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
  <footer>MediBook Selenium E2E Suite &mdash; Phase 7 CI/CD</footer>
</body>
</html>"""


def generate_html(report: dict, base_url: str) -> str:
    if not HAS_JINJA2:
        print("[WARN] jinja2 not installed — writing raw HTML.")
        out_path = str(HTML_DIR / "execution-report.html")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(f"<html><body><h1>MediBook Test Report</h1>"
                    f"<p>Total: {report['total']} Passed: {report['passed']} "
                    f"Failed: {report['failed']}</p></body></html>")
        return out_path

    pct = round(report["passed"] / report["total"] * 100, 1) if report["total"] else 0
    env = Environment(loader=BaseLoader())
    tmpl = env.from_string(HTML_TEMPLATE)
    rendered = tmpl.render(**report, base_url=base_url, pct=pct)
    out_path = str(HTML_DIR / "execution-report.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(rendered)
    print(f"[OK] HTML report -> {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────
# Markdown summary
# ─────────────────────────────────────────────────────────────
def generate_summary(report: dict, base_url: str) -> str:
    pct = round(report["passed"] / report["total"] * 100, 1) if report["total"] else 0
    failed = [t for t in report["tests"] if t["outcome"] == "failed"]

    lines = [
        "# Live GitHub Pages E2E Test Summary\n",
        f"**Deployment URL:** {base_url}  ",
        f"**Generated At:** {report['generated_at']}  ",
        f"**Total Duration:** {report['duration']}s\n",
        "## Results\n",
        f"| Metric | Value |",
        f"|--------|-------|",
        f"| Total Tests | {report['total']} |",
        f"| Passed | ✅ {report['passed']} |",
        f"| Failed | ❌ {report['failed']} |",
        f"| Skipped | ⏭️ {report['skipped']} |",
        f"| Pass Rate | **{pct}%** |\n",
    ]

    if failed:
        lines += ["\n## Failed Tests\n"]
        for t in failed:
            lines.append(f"- **{t['name']}** — {str(t['error'])[:200] if t['error'] else 'assertion error'}")
    else:
        lines.append("\n✅ All tests passed!")

    out_path = str(SUMMARY_DIR / "summary.md")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"[OK] Summary -> {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────
# Fallback: create empty report if no JSON exists
# ─────────────────────────────────────────────────────────────
def empty_report(reason: str = "No test data") -> dict:
    return {
        "tests": [], "total": 0, "passed": 0,
        "failed": 0, "skipped": 0, "duration": 0,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate MediBook test reports.")
    parser.add_argument("--json",     default="test-results.json", help="Path to pytest JSON report")
    parser.add_argument("--base-url", default="http://localhost:5173", help="Deployed URL")
    args = parser.parse_args()

    base_url = args.base_url.rstrip("/")

    if os.path.exists(args.json):
        report = parse_json_report(args.json)
    else:
        print(f"[WARN] JSON report not found at {args.json} — using empty report.")
        report = empty_report()

    generate_excel(report, base_url)
    generate_html(report, base_url)
    generate_summary(report, base_url)

    print(f"\n[DONE] Reports generated in: {RESULTS_DIR}")


if __name__ == "__main__":
    main()
