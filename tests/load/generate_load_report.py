"""
generate_load_report.py
========================
Reads Locust CSV stats files and produces:
  1. A rich dark-themed HTML report (with charts via Chart.js)
  2. An Excel workbook (colour-coded per endpoint)
  3. A Markdown summary

Usage (called by run_load_test.py automatically):
    python tests/load/generate_load_report.py \
        --csv-prefix  tests/load/results/stats \
        --users       100 \
        --duration    60 \
        --host        http://localhost:5000

Output:
    tests/load/results/
    ├── load-test-report.html
    ├── Load_Test_Report.xlsx
    └── summary.md
"""

import argparse
import csv
import json
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from jinja2 import Environment, BaseLoader
    HAS_JINJA2 = True
except ImportError:
    HAS_JINJA2 = False

RESULTS_DIR = Path(__file__).parent / "results"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)


# ─────────────────────────────────────────────────────────────
# Parse Locust CSV stats
# ─────────────────────────────────────────────────────────────
def parse_locust_csv(csv_prefix: str) -> dict:
    """
    Locust produces two CSV files:
      <prefix>_stats.csv      — per-endpoint aggregated stats
      <prefix>_stats_history.csv — time-series RPS/response-time data
    """
    stats_file   = f"{csv_prefix}_stats.csv"
    history_file = f"{csv_prefix}_stats_history.csv"

    endpoints  = []
    total_row  = None
    rps_series = []  # [{time, rps, avg_rt}]

    # ── Parse _stats.csv ─────────────────────────────────────
    if os.path.exists(stats_file):
        with open(stats_file, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                entry = {
                    "name":         row.get("Name", ""),
                    "method":       row.get("Type", ""),
                    "requests":     int(row.get("Request Count", 0)),
                    "failures":     int(row.get("Failure Count", 0)),
                    "avg_rt":       float(row.get("Average Response Time", 0)),
                    "min_rt":       float(row.get("Min Response Time", 0)),
                    "max_rt":       float(row.get("Max Response Time", 0)),
                    "p50":          float(row.get("50%", 0)),
                    "p90":          float(row.get("90%", 0)),
                    "p95":          float(row.get("95%", 0)),
                    "p99":          float(row.get("99%", 0)),
                    "rps":          float(row.get("Requests/s", 0)),
                    "fail_rate":    0,
                }
                reqs = entry["requests"]
                if reqs > 0:
                    entry["fail_rate"] = round(entry["failures"] / reqs * 100, 1)

                if row.get("Name") in ("Aggregated", "Total"):
                    total_row = entry
                else:
                    endpoints.append(entry)

    # ── Parse _stats_history.csv ──────────────────────────────
    if os.path.exists(history_file):
        with open(history_file, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get("Name", "")
                if name in ("Aggregated", "Total", ""):
                    try:
                        rps_series.append({
                            "timestamp": float(row.get("Timestamp", 0)),
                            "users":     int(float(row.get("User Count", 0))),
                            "rps":       float(row.get("Requests/s", 0)),
                            "failures":  float(row.get("Failures/s", 0)),
                            "avg_rt":    float(row.get("Total Average Response Time", 0)),
                            "p50":       float(row.get("50%", 0)),
                            "p95":       float(row.get("95%", 0)),
                        })
                    except (ValueError, TypeError):
                        pass

    # ── Build totals if CSV not found ─────────────────────────
    if total_row is None:
        all_reqs  = sum(e["requests"] for e in endpoints)
        all_fails = sum(e["failures"] for e in endpoints)
        total_row = {
            "name":      "Aggregated",
            "requests":  all_reqs,
            "failures":  all_fails,
            "avg_rt":    sum(e["avg_rt"] for e in endpoints) / len(endpoints) if endpoints else 0,
            "min_rt":    min((e["min_rt"] for e in endpoints), default=0),
            "max_rt":    max((e["max_rt"] for e in endpoints), default=0),
            "p50":       sum(e["p50"] for e in endpoints) / len(endpoints) if endpoints else 0,
            "p90":       sum(e["p90"] for e in endpoints) / len(endpoints) if endpoints else 0,
            "p99":       sum(e["p99"] for e in endpoints) / len(endpoints) if endpoints else 0,
            "rps":       sum(e["rps"] for e in endpoints),
            "fail_rate": round(all_fails / all_reqs * 100, 1) if all_reqs else 0,
        }

    return {
        "endpoints":   endpoints,
        "total":       total_row,
        "rps_series":  rps_series,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _status(avg_rt: float, fail_rate: float) -> str:
    if fail_rate > 5:
        return "FAIL"
    if avg_rt < 300:
        return "EXCELLENT"
    if avg_rt < 800:
        return "GOOD"
    if avg_rt < 2000:
        return "ACCEPTABLE"
    return "SLOW"


# ─────────────────────────────────────────────────────────────
# HTML Report
# ─────────────────────────────────────────────────────────────
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1.0"/>
  <title>MediBook Load Test Report</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0a0a14;color:#e2e8f0;padding:2rem}
    h1{font-size:2rem;background:linear-gradient(135deg,#7c3aed,#2563eb);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:.25rem}
    h2{font-size:1.1rem;color:#94a3b8;font-weight:600;text-transform:uppercase;letter-spacing:.08em;margin:2rem 0 1rem}
    .subtitle{color:#64748b;font-size:.9rem;margin-bottom:2rem}
    .config-bar{display:flex;flex-wrap:wrap;gap:1.5rem;background:#13131f;border:1px solid #1e293b;border-radius:10px;padding:1rem 1.5rem;margin-bottom:2rem;font-size:.85rem}
    .config-bar .item span{color:#64748b}.config-bar .item strong{color:#a5b4fc;margin-left:.4rem}
    .kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:1rem;margin-bottom:2rem}
    .kpi{background:#13131f;border-radius:14px;padding:1.5rem 1.25rem;text-align:center;border:1px solid #1e293b;position:relative;overflow:hidden}
    .kpi::before{content:'';position:absolute;inset:0;background:var(--glow);opacity:.06;border-radius:inherit}
    .kpi.rps{--glow:linear-gradient(135deg,#7c3aed,#2563eb)}
    .kpi.avg{--glow:linear-gradient(135deg,#10b981,#34d399)}
    .kpi.min{--glow:linear-gradient(135deg,#06b6d4,#67e8f9)}
    .kpi.max{--glow:linear-gradient(135deg,#f59e0b,#fcd34d)}
    .kpi.p90{--glow:linear-gradient(135deg,#8b5cf6,#a78bfa)}
    .kpi.p99{--glow:linear-gradient(135deg,#ef4444,#fca5a5)}
    .kpi.err{--glow:linear-gradient(135deg,#ef4444,#dc2626)}
    .kpi .val{font-size:2.2rem;font-weight:800;line-height:1}
    .kpi.rps .val{color:#a5b4fc}.kpi.avg .val{color:#10b981}.kpi.min .val{color:#67e8f9}
    .kpi.max .val{color:#fcd34d}.kpi.p90 .val{color:#a78bfa}.kpi.p99 .val{color:#fca5a5}
    .kpi.err .val{color:#ef4444}
    .kpi .unit{font-size:.8rem;color:#64748b;margin-top:.15rem}
    .kpi .lbl{font-size:.7rem;text-transform:uppercase;letter-spacing:.1em;color:#475569;margin-top:.5rem}
    .charts-grid{display:grid;grid-template-columns:1fr 1fr;gap:1.5rem;margin-bottom:2rem}
    @media(max-width:900px){.charts-grid{grid-template-columns:1fr}}
    .chart-card{background:#13131f;border-radius:14px;padding:1.5rem;border:1px solid #1e293b}
    .chart-card h3{font-size:.85rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.08em;margin-bottom:1rem}
    table{width:100%;border-collapse:collapse;background:#13131f;border-radius:12px;overflow:hidden;margin-bottom:2rem}
    th{background:#1e1b4b;padding:.75rem 1rem;text-align:left;font-size:.7rem;text-transform:uppercase;letter-spacing:.05em;color:#64748b}
    td{padding:.75rem 1rem;border-bottom:1px solid #1a1a2e;font-size:.85rem}
    tr:last-child td{border-bottom:none}tr:hover td{background:#0f0f1f}
    .badge{display:inline-block;padding:.2rem .65rem;border-radius:9999px;font-size:.7rem;font-weight:700}
    .badge.EXCELLENT{background:#022c22;color:#10b981}
    .badge.GOOD{background:#052e16;color:#34d399}
    .badge.ACCEPTABLE{background:#451a03;color:#f59e0b}
    .badge.SLOW{background:#450a0a;color:#ef4444}
    .badge.FAIL{background:#7f1d1d;color:#fca5a5}
    .rt-bar{height:6px;border-radius:9999px;background:linear-gradient(90deg,#10b981,#7c3aed);margin-top:4px}
    footer{text-align:center;color:#334155;font-size:.8rem;margin-top:2rem;border-top:1px solid #1e293b;padding-top:1rem}
  </style>
</head>
<body>
  <h1>MediBook — Load Test Report</h1>
  <p class="subtitle">Baseline Load Test · {{ users }} Virtual Users · {{ duration }}s Duration · {{ generated_at }}</p>

  <div class="config-bar">
    <div class="item"><span>Host</span><strong>{{ host }}</strong></div>
    <div class="item"><span>Virtual Users</span><strong>{{ users }}</strong></div>
    <div class="item"><span>Duration</span><strong>{{ duration }}s</strong></div>
    <div class="item"><span>Spawn Rate</span><strong>10 users/s</strong></div>
    <div class="item"><span>Total Requests</span><strong>{{ total.requests }}</strong></div>
    <div class="item"><span>Failures</span><strong style="color:{% if total.fail_rate > 5 %}#ef4444{% else %}#10b981{% endif %}">{{ total.failures }}</strong></div>
  </div>

  <!-- KPI Cards -->
  <div class="kpi-grid">
    <div class="kpi rps">
      <div class="val">{{ "%.1f"|format(total.rps) }}</div>
      <div class="unit">req/sec</div>
      <div class="lbl">Throughput</div>
    </div>
    <div class="kpi avg">
      <div class="val">{{ total.avg_rt|int }}</div>
      <div class="unit">ms</div>
      <div class="lbl">Avg Response</div>
    </div>
    <div class="kpi min">
      <div class="val">{{ total.min_rt|int }}</div>
      <div class="unit">ms</div>
      <div class="lbl">Min Response</div>
    </div>
    <div class="kpi max">
      <div class="val">{{ total.max_rt|int }}</div>
      <div class="unit">ms</div>
      <div class="lbl">Max Response</div>
    </div>
    <div class="kpi p90">
      <div class="val">{{ total.p90|int }}</div>
      <div class="unit">ms</div>
      <div class="lbl">P90 Response</div>
    </div>
    <div class="kpi p99">
      <div class="val">{{ total.p99|int }}</div>
      <div class="unit">ms</div>
      <div class="lbl">P99 Response</div>
    </div>
    <div class="kpi err">
      <div class="val">{{ total.fail_rate }}</div>
      <div class="unit">%</div>
      <div class="lbl">Error Rate</div>
    </div>
  </div>

  <!-- Charts -->
  <div class="charts-grid">
    <div class="chart-card">
      <h3>📈 Requests per Second over Time</h3>
      <canvas id="rpsChart" height="200"></canvas>
    </div>
    <div class="chart-card">
      <h3>⏱ Response Time over Time (ms)</h3>
      <canvas id="rtChart" height="200"></canvas>
    </div>
    <div class="chart-card">
      <h3>🔀 Response Time by Endpoint (ms)</h3>
      <canvas id="endpointChart" height="200"></canvas>
    </div>
    <div class="chart-card">
      <h3>👥 Active Users over Time</h3>
      <canvas id="usersChart" height="200"></canvas>
    </div>
  </div>

  <!-- Endpoint table -->
  <h2>Per-Endpoint Breakdown</h2>
  <table>
    <thead>
      <tr>
        <th>Endpoint</th><th>Method</th><th>Requests</th><th>Failures</th>
        <th>Avg (ms)</th><th>Min (ms)</th><th>Max (ms)</th>
        <th>P50</th><th>P90</th><th>P99</th><th>RPS</th><th>Status</th>
      </tr>
    </thead>
    <tbody>
      {% for ep in endpoints %}
      <tr>
        <td><code style="color:#a5b4fc;font-size:.8rem">{{ ep.name }}</code></td>
        <td><span style="color:#64748b;font-size:.75rem">{{ ep.method }}</span></td>
        <td>{{ ep.requests }}</td>
        <td style="color:{% if ep.failures > 0 %}#ef4444{% else %}#10b981{% endif %}">{{ ep.failures }}</td>
        <td>
          {{ ep.avg_rt|int }}
          <div class="rt-bar" style="width:{% if total.max_rt > 0 %}{{ [ep.avg_rt/total.max_rt*100,100]|min|int }}{% else %}0{% endif %}%"></div>
        </td>
        <td style="color:#67e8f9">{{ ep.min_rt|int }}</td>
        <td style="color:#fcd34d">{{ ep.max_rt|int }}</td>
        <td>{{ ep.p50|int }}</td>
        <td>{{ ep.p90|int }}</td>
        <td>{{ ep.p99|int }}</td>
        <td style="color:#a5b4fc">{{ "%.1f"|format(ep.rps) }}</td>
        <td><span class="badge {{ ep.status }}">{{ ep.status }}</span></td>
      </tr>
      {% endfor %}
      <tr style="background:#1e1b4b">
        <td colspan="2"><strong style="color:#a5b4fc">TOTAL</strong></td>
        <td><strong>{{ total.requests }}</strong></td>
        <td><strong style="color:{% if total.failures > 0 %}#ef4444{% else %}#10b981{% endif %}">{{ total.failures }}</strong></td>
        <td><strong>{{ total.avg_rt|int }}</strong></td>
        <td><strong style="color:#67e8f9">{{ total.min_rt|int }}</strong></td>
        <td><strong style="color:#fcd34d">{{ total.max_rt|int }}</strong></td>
        <td><strong>{{ total.p50|int }}</strong></td>
        <td><strong>{{ total.p90|int }}</strong></td>
        <td><strong>{{ total.p99|int }}</strong></td>
        <td><strong style="color:#a5b4fc">{{ "%.1f"|format(total.rps) }}</strong></td>
        <td></td>
      </tr>
    </tbody>
  </table>

  <footer>MediBook Baseline Load Test · Locust {{ users }} VU · {{ duration }}s</footer>

<script>
const PURPLE = '#7c3aed', GREEN = '#10b981', AMBER = '#f59e0b', BLUE = '#3b82f6';
const chartDefaults = {
  responsive: true,
  plugins: { legend: { labels: { color: '#94a3b8', font: { size: 11 } } } },
  scales: {
    x: { ticks: { color: '#475569', maxTicksLimit: 10 }, grid: { color: '#1e293b' } },
    y: { ticks: { color: '#475569' }, grid: { color: '#1e293b' } }
  }
};

const rpsData    = {{ rps_labels | tojson }};
const rpsValues  = {{ rps_values | tojson }};
const failValues = {{ fail_values | tojson }};
const rtAvg      = {{ rt_avg | tojson }};
const rtP95      = {{ rt_p95 | tojson }};
const userCounts = {{ user_counts | tojson }};
const epNames    = {{ ep_names | tojson }};
const epAvgRt    = {{ ep_avg_rt | tojson }};
const epP90      = {{ ep_p90 | tojson }};

new Chart(document.getElementById('rpsChart'), {
  type: 'line',
  data: {
    labels: rpsData,
    datasets: [
      { label: 'RPS', data: rpsValues, borderColor: PURPLE, backgroundColor: PURPLE+'22', fill:true, tension:.3, pointRadius:0 },
      { label: 'Failures/s', data: failValues, borderColor: '#ef4444', backgroundColor: '#ef444422', fill:true, tension:.3, pointRadius:0 },
    ]
  },
  options: { ...chartDefaults }
});

new Chart(document.getElementById('rtChart'), {
  type: 'line',
  data: {
    labels: rpsData,
    datasets: [
      { label: 'Avg (ms)',  data: rtAvg, borderColor: GREEN,  backgroundColor: GREEN+'22',  fill:true, tension:.3, pointRadius:0 },
      { label: 'P95 (ms)',  data: rtP95, borderColor: AMBER,  backgroundColor: AMBER+'22',  fill:true, tension:.3, pointRadius:0 },
    ]
  },
  options: { ...chartDefaults }
});

new Chart(document.getElementById('endpointChart'), {
  type: 'bar',
  data: {
    labels: epNames,
    datasets: [
      { label: 'Avg RT (ms)', data: epAvgRt, backgroundColor: PURPLE+'bb' },
      { label: 'P90 RT (ms)', data: epP90,   backgroundColor: AMBER+'bb'  },
    ]
  },
  options: { ...chartDefaults, indexAxis: 'y' }
});

new Chart(document.getElementById('usersChart'), {
  type: 'line',
  data: {
    labels: rpsData,
    datasets: [
      { label: 'Active Users', data: userCounts, borderColor: BLUE, backgroundColor: BLUE+'22', fill:true, tension:.3, pointRadius:0 },
    ]
  },
  options: { ...chartDefaults }
});
</script>
</body>
</html>"""


def generate_html(data: dict, users: int, duration: int, host: str) -> str:
    if not HAS_JINJA2:
        out = str(RESULTS_DIR / "load-test-report.html")
        with open(out, "w", encoding="utf-8") as f:
            t = data["total"]
            f.write(f"<html><body><h1>Load Test</h1>"
                    f"<p>RPS: {t['rps']:.1f} | Avg: {t['avg_rt']:.0f}ms | "
                    f"Min: {t['min_rt']:.0f}ms | Max: {t['max_rt']:.0f}ms</p></body></html>")
        return out

    # Prepare chart series from rps_series
    series = data["rps_series"]
    t0     = series[0]["timestamp"] if series else 0
    labels      = [f"{int(s['timestamp'] - t0)}s" for s in series]
    rps_vals    = [round(s["rps"], 1)      for s in series]
    fail_vals   = [round(s["failures"], 1) for s in series]
    rt_avg      = [round(s["avg_rt"], 0)   for s in series]
    rt_p95      = [round(s["p95"], 0)      for s in series]
    user_counts = [s["users"]              for s in series]

    eps     = data["endpoints"]
    ep_names  = [e["name"] for e in eps]
    ep_avg_rt = [round(e["avg_rt"], 0) for e in eps]
    ep_p90    = [round(e["p90"], 0)    for e in eps]

    # Add status to each endpoint
    for ep in eps:
        ep["status"] = _status(ep["avg_rt"], ep["fail_rate"])

    env  = Environment(loader=BaseLoader())
    tmpl = env.from_string(HTML_TEMPLATE)
    html = tmpl.render(
        **data,
        users=users, duration=duration, host=host,
        rps_labels=labels, rps_values=rps_vals, fail_values=fail_vals,
        rt_avg=rt_avg, rt_p95=rt_p95, user_counts=user_counts,
        ep_names=ep_names, ep_avg_rt=ep_avg_rt, ep_p90=ep_p90,
    )
    out = str(RESULTS_DIR / "load-test-report.html")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] HTML report saved: {out}")
    return out


# ─────────────────────────────────────────────────────────────
# Excel Report
# ─────────────────────────────────────────────────────────────
def generate_excel(data: dict, users: int, duration: int, host: str) -> str:
    if not HAS_OPENPYXL:
        print("[WARN] openpyxl not installed.")
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    PURPLE = "7C3AED"; GREEN  = "10B981"; RED   = "EF4444"
    AMBER  = "F59E0B"; DARK   = "1E1B4B"; WHITE = "FFFFFF"; LGRAY = "F1F5F9"

    # Title
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = "MediBook — Baseline Load Test Report"
    c.font  = Font(bold=True, size=16, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    t = data["total"]
    meta = [
        ("Host",            host),
        ("Virtual Users",   users),
        ("Duration",        f"{duration} seconds"),
        ("Generated At",    data["generated_at"]),
        ("Total Requests",  t["requests"]),
        ("Total Failures",  t["failures"]),
    ]
    for i, (lbl, val) in enumerate(meta, 2):
        ws.cell(i, 1, lbl).font = Font(bold=True, color=DARK)
        ws.cell(i, 2, val)

    # KPI section
    ws.merge_cells("A9:L9")
    ws["A9"].value = "PERFORMANCE METRICS"
    ws["A9"].font  = Font(bold=True, size=12, color=WHITE)
    ws["A9"].fill  = PatternFill("solid", fgColor=PURPLE)
    ws["A9"].alignment = Alignment(horizontal="center")

    kpis = [
        ("RPS",       f"{t['rps']:.1f} req/s",  "4F46E5"),
        ("Avg RT",    f"{t['avg_rt']:.0f} ms",   GREEN),
        ("Min RT",    f"{t['min_rt']:.0f} ms",   "06B6D4"),
        ("Max RT",    f"{t['max_rt']:.0f} ms",   AMBER),
        ("P90",       f"{t['p90']:.0f} ms",      "8B5CF6"),
        ("P99",       f"{t['p99']:.0f} ms",      "EC4899"),
        ("Err Rate",  f"{t['fail_rate']}%",       RED if t["fail_rate"] > 1 else GREEN),
    ]
    for col, (lbl, val, clr) in enumerate(kpis, 1):
        ws.cell(10, col, lbl).font  = Font(bold=True, color=WHITE)
        ws.cell(10, col, lbl).fill  = PatternFill("solid", fgColor=clr)
        ws.cell(10, col, lbl).alignment = Alignment(horizontal="center")
        ws.cell(11, col, val).font  = Font(bold=True, size=13, color=clr)
        ws.cell(11, col, val).alignment = Alignment(horizontal="center")
    ws.row_dimensions[11].height = 30

    # Endpoint detail sheet
    ws2 = wb.create_sheet("Endpoint Details")
    hdr = ["#", "Endpoint", "Method", "Requests", "Failures", "Fail%",
           "Avg RT (ms)", "Min RT (ms)", "Max RT (ms)", "P50", "P90", "P99", "RPS", "Status"]
    for col, h in enumerate(hdr, 1):
        cell = ws2.cell(1, col, h)
        cell.font  = Font(bold=True, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="center")

    STATUS_COLOUR = {"EXCELLENT": GREEN, "GOOD": "34D399",
                     "ACCEPTABLE": AMBER, "SLOW": RED, "FAIL": "7F1D1D"}
    for row, ep in enumerate(data["endpoints"], 2):
        status = _status(ep["avg_rt"], ep["fail_rate"])
        sc     = STATUS_COLOUR.get(status, "888888")
        vals   = [row-1, ep["name"], ep.get("method",""), ep["requests"], ep["failures"],
                  f"{ep['fail_rate']}%", round(ep["avg_rt"]), round(ep["min_rt"]),
                  round(ep["max_rt"]), round(ep["p50"]), round(ep["p90"]),
                  round(ep["p99"]), round(ep["rps"], 1), status]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row, col, v)
            if col == 14:
                cell.font = Font(bold=True, color=sc)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LGRAY)

    for sheet in [ws, ws2]:
        for col in sheet.columns:
            w = max((len(str(c.value or "")) for c in col), default=0)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    out = str(RESULTS_DIR / "Load_Test_Report.xlsx")
    wb.save(out)
    print(f"[OK] Excel -> {out}")
    return out


# ─────────────────────────────────────────────────────────────
# Markdown Summary
# ─────────────────────────────────────────────────────────────
def generate_summary(data: dict, users: int, duration: int, host: str) -> str:
    t   = data["total"]
    pct = round((1 - t["fail_rate"] / 100) * 100, 1) if t["requests"] else 0
    verdict = _status(t["avg_rt"], t["fail_rate"])

    lines = [
        "# MediBook Baseline Load Test — Summary\n",
        f"**Host:** {host}  ",
        f"**Generated At:** {data['generated_at']}  \n",
        "## Configuration\n",
        "| Setting | Value |", "|---------|-------|",
        f"| Virtual Users | **{users}** |",
        f"| Duration | **{duration} seconds** |",
        f"| Spawn Rate | 10 users/s |",
        f"| Total Requests | {t['requests']} |\n",
        "## Results\n",
        "| Metric | Value |", "|--------|-------|",
        f"| Requests/sec (RPS) | **{t['rps']:.1f} req/sec** |",
        f"| Average Response Time | **{t['avg_rt']:.0f} ms** |",
        f"| Min Response Time | {t['min_rt']:.0f} ms |",
        f"| Max Response Time | {t['max_rt']:.0f} ms |",
        f"| P50 (median) | {t['p50']:.0f} ms |",
        f"| P90 | {t['p90']:.0f} ms |",
        f"| P99 | {t['p99']:.0f} ms |",
        f"| Failures | {t['failures']} ({t['fail_rate']}%) |",
        f"| Overall Status | **{verdict}** |\n",
        "## Endpoint Breakdown\n",
        "| Endpoint | Avg (ms) | Min | Max | P90 | RPS | Status |",
        "|----------|----------|-----|-----|-----|-----|--------|",
    ]
    for ep in data["endpoints"]:
        st = _status(ep["avg_rt"], ep["fail_rate"])
        lines.append(
            f"| `{ep['name']}` | {ep['avg_rt']:.0f} | {ep['min_rt']:.0f} | "
            f"{ep['max_rt']:.0f} | {ep['p90']:.0f} | {ep['rps']:.1f} | {st} |"
        )

    out = str(RESULTS_DIR / "summary.md")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"[OK] Summary -> {out}")
    return out


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv-prefix", default="tests/load/results/stats")
    parser.add_argument("--users",      type=int, default=100)
    parser.add_argument("--duration",   type=int, default=60)
    parser.add_argument("--host",       default="http://localhost:5000")
    args = parser.parse_args()

    print(f"\n[load-report] Generating reports from: {args.csv_prefix}_stats.csv")
    data = parse_locust_csv(args.csv_prefix)
    generate_html(data,    args.users, args.duration, args.host)
    generate_excel(data,   args.users, args.duration, args.host)
    generate_summary(data, args.users, args.duration, args.host)
    print(f"\n[DONE] Reports saved to: {RESULTS_DIR}")


if __name__ == "__main__":
    main()
