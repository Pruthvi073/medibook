"""
generate_report.py — Generates comprehensive Excel analysis reports, HTML dashboards,
and Markdown execution summaries for MediBook Appium Mobile E2E automation.
"""
import argparse
import json
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from jinja2 import Environment, BaseLoader
    HAS_JINJA2 = True
except ImportError:
    HAS_JINJA2 = False

# Setup directories
ROOT        = Path(__file__).parent.parent
RESULTS_DIR = ROOT / "Test Results"
EXCEL_DIR   = RESULTS_DIR / "Excel"
HTML_DIR    = RESULTS_DIR / "HTML"
SUMMARY_DIR = RESULTS_DIR / "Summary"

for d in [EXCEL_DIR, HTML_DIR, SUMMARY_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ─────────────────────────────────────────────────────────────
# Parse JSON report
# ─────────────────────────────────────────────────────────────
def parse_results(json_path: str) -> dict:
    # Load database test cases for rich metadata
    db_path = ROOT / "data" / "test_cases.json"
    tc_db = {}
    if db_path.exists():
        with open(db_path, "r", encoding="utf-8") as f:
            for tc in json.load(f):
                tc_db[tc["id"]] = tc

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    tests = []
    for t in data.get("tests", []):
        nodeid = t.get("nodeid", "")
        # Extract parameter ID (bracket value, e.g., TC_AUTH_001)
        tc_id = "UNKNOWN"
        if "[" in nodeid and "]" in nodeid:
            tc_id = nodeid.split("[")[-1].rstrip("]")

        outcome = t.get("outcome", "unknown")
        duration = round(t.get("call", {}).get("duration", 0) + t.get("setup", {}).get("duration", 0), 2)
        
        # Pull details from DB
        meta = tc_db.get(tc_id, {
            "module": "Unknown",
            "name": f"Dynamic Test Case {tc_id}",
            "priority": "Medium",
            "steps": "1. Run automated steps",
            "expected": "Success outcome",
            "error_reason": ""
        })

        err_msg = ""
        if outcome == "failed":
            err_msg = t.get("call", {}).get("longrepr", "")
            if not err_msg:
                err_msg = "AssertionError: Expected outcome mismatched."
        elif outcome == "skipped":
            err_msg = t.get("setup", {}).get("longrepr", "")
            if not err_msg:
                err_msg = meta.get("error_reason", "Feature Disabled")

        tests.append({
            "id": tc_id,
            "module": meta["module"],
            "name": meta["name"],
            "priority": meta["priority"],
            "steps": meta["steps"],
            "expected": meta["expected"],
            "outcome": outcome,
            "duration": duration,
            "error": err_msg[:400] if err_msg else ""
        })

    s = data.get("summary", {})
    total = s.get("total", len(tests))
    passed = s.get("passed", sum(1 for t in tests if t["outcome"] == "passed"))
    failed = s.get("failed", sum(1 for t in tests if t["outcome"] == "failed"))
    skipped = s.get("skipped", sum(1 for t in tests if t["outcome"] == "skipped"))

    return {
        "tests": tests,
        "total": total,
        "passed": passed,
        "failed": failed,
        "skipped": skipped,
        "blocked": 0,  # Blocked is a subclass of skipped/failed in this harness
        "duration": round(data.get("duration", 0), 2),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

# ─────────────────────────────────────────────────────────────
# Excel Reports Generation
# ─────────────────────────────────────────────────────────────
def generate_excel_reports(report: dict, base_url: str, run_number: str):
    if not HAS_OPENPYXL:
        print("[WARN] openpyxl not installed - skipping Excel generation.")
        return

    # Color definitions
    DARK   = "1E1B4B"; WHITE = "FFFFFF"; LGRAY = "F1F5F9"
    PURPLE = "7C3AED"; GREEN = "10B981"; RED   = "EF4444"; AMBER  = "F59E0B"

    # -- Report 1: Automation_Test_Report.xlsx (Enterprise-grade) ----------------
    wb = openpyxl.Workbook()
    
    # Sheet 1: Executed Test Cases
    ws1 = wb.active
    ws1.title = "Executed Test Cases"
    headers1 = ["Test ID", "Module", "Test Name", "Priority", "Status", "Execution Time (s)"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(1, col, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="center")

    for row, t in enumerate(report["tests"], 2):
        ws1.cell(row, 1, t["id"])
        ws1.cell(row, 2, t["module"])
        ws1.cell(row, 3, t["name"])
        ws1.cell(row, 4, t["priority"])
        
        status_cell = ws1.cell(row, 5, t["outcome"].upper())
        status_cell.font = Font(bold=True, color=GREEN if t["outcome"] == "passed" else RED if t["outcome"] == "failed" else AMBER)
        
        ws1.cell(row, 6, t["duration"])
        if row % 2 == 0:
            for c in range(1, 7):
                ws1.cell(row, c).fill = PatternFill("solid", fgColor=LGRAY)

    # Sheet 2: Passed Tests
    ws2 = wb.create_sheet("Passed Tests")
    for col, h in enumerate(headers1, 1):
        cell = ws2.cell(1, col, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.alignment = Alignment(horizontal="center")
    
    pass_row = 2
    for t in report["tests"]:
        if t["outcome"] == "passed":
            ws2.cell(pass_row, 1, t["id"])
            ws2.cell(pass_row, 2, t["module"])
            ws2.cell(pass_row, 3, t["name"])
            ws2.cell(pass_row, 4, t["priority"])
            ws2.cell(pass_row, 5, "PASSED").font = Font(bold=True, color=GREEN)
            ws2.cell(pass_row, 6, t["duration"])
            pass_row += 1

    # Sheet 3: Failed Tests
    ws3 = wb.create_sheet("Failed Tests")
    headers_fail = headers1 + ["Failure Reason"]
    for col, h in enumerate(headers_fail, 1):
        cell = ws3.cell(1, col, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=RED)
        cell.alignment = Alignment(horizontal="center")

    fail_row = 2
    for t in report["tests"]:
        if t["outcome"] == "failed":
            ws3.cell(fail_row, 1, t["id"])
            ws3.cell(fail_row, 2, t["module"])
            ws3.cell(fail_row, 3, t["name"])
            ws3.cell(fail_row, 4, t["priority"])
            ws3.cell(fail_row, 5, "FAILED").font = Font(bold=True, color=RED)
            ws3.cell(fail_row, 6, t["duration"])
            ws3.cell(fail_row, 7, t["error"])
            fail_row += 1

    # Sheet 4: Skipped Tests
    ws4 = wb.create_sheet("Skipped Tests")
    headers_skip = headers1 + ["Skip Reason"]
    for col, h in enumerate(headers_skip, 1):
        cell = ws4.cell(1, col, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=AMBER)
        cell.alignment = Alignment(horizontal="center")

    skip_row = 2
    for t in report["tests"]:
        if t["outcome"] == "skipped":
            ws4.cell(skip_row, 1, t["id"])
            ws4.cell(skip_row, 2, t["module"])
            ws4.cell(skip_row, 3, t["name"])
            ws4.cell(skip_row, 4, t["priority"])
            ws4.cell(skip_row, 5, "SKIPPED").font = Font(bold=True, color=AMBER)
            ws4.cell(skip_row, 6, t["duration"])
            ws4.cell(skip_row, 7, t["error"])
            skip_row += 1

    # Sheet 5: Execution Metrics
    ws5 = wb.create_sheet("Execution Metrics")
    ws5.cell(1, 1, "Metric").font = Font(bold=True, color=WHITE)
    ws5.cell(1, 1).fill = PatternFill("solid", fgColor=DARK)
    ws5.cell(1, 2, "Value").font = Font(bold=True, color=WHITE)
    ws5.cell(1, 2).fill = PatternFill("solid", fgColor=DARK)
    
    metrics = [
        ("Total Tests",    report["total"]),
        ("Passed Tests",   report["passed"]),
        ("Failed Tests",   report["failed"]),
        ("Skipped Tests",  report["skipped"]),
        ("Pass Rate %",    round(report["passed"] / report["total"] * 100, 2) if report["total"] else 0),
        ("Execution Time", f"{report['duration']} s"),
    ]
    for r, (m, v) in enumerate(metrics, 2):
        ws5.cell(r, 1, m).font = Font(bold=True)
        ws5.cell(r, 2, v)

    # Sheet 6: Defect Summary
    ws6 = wb.create_sheet("Defect Summary")
    ws6.cell(1, 1, "Module").font = Font(bold=True, color=WHITE)
    ws6.cell(1, 1).fill = PatternFill("solid", fgColor=RED)
    ws6.cell(1, 2, "Failed Test Count").font = Font(bold=True, color=WHITE)
    ws6.cell(1, 2).fill = PatternFill("solid", fgColor=RED)
    
    module_defects = {}
    for t in report["tests"]:
        if t["outcome"] == "failed":
            module_defects[t["module"]] = module_defects.get(t["module"], 0) + 1
            
    for r, (mod, count) in enumerate(module_defects.items(), 2):
        ws6.cell(r, 1, mod)
        ws6.cell(r, 2, count)

    # Sheet 7: Pass Rate Summary
    ws7 = wb.create_sheet("Pass Rate Summary")
    ws7.cell(1, 1, "Module").font = Font(bold=True, color=WHITE)
    ws7.cell(1, 1).fill = PatternFill("solid", fgColor=PURPLE)
    ws7.cell(1, 2, "Total").font = Font(bold=True, color=WHITE)
    ws7.cell(1, 2).fill = PatternFill("solid", fgColor=PURPLE)
    ws7.cell(1, 3, "Passed").font = Font(bold=True, color=WHITE)
    ws7.cell(1, 3).fill = PatternFill("solid", fgColor=PURPLE)
    ws7.cell(1, 4, "Pass Rate %").font = Font(bold=True, color=WHITE)
    ws7.cell(1, 4).fill = PatternFill("solid", fgColor=PURPLE)

    module_stats = {}
    for t in report["tests"]:
        m = t["module"]
        if m not in module_stats:
            module_stats[m] = {"total": 0, "passed": 0}
        module_stats[m]["total"] += 1
        if t["outcome"] == "passed":
            module_stats[m]["passed"] += 1

    for r, (mod, s) in enumerate(module_stats.items(), 2):
        ws7.cell(r, 1, mod)
        ws7.cell(r, 2, s["total"])
        ws7.cell(r, 3, s["passed"])
        rate = round(s["passed"] / s["total"] * 100, 1) if s["total"] else 0
        ws7.cell(r, 4, f"{rate}%")

    # Save Auto-widths for all sheets
    for ws_item in [ws1, ws2, ws3, ws4, ws5, ws6, ws7]:
        for col in ws_item.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws_item.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    report_path = str(EXCEL_DIR / "Automation_Test_Report.xlsx")
    wb.save(report_path)
    print(f"[OK] Excel report -> {report_path}")

    # -- Report 2: Passed_Test_Cases.xlsx ----------------------------------------
    wb_pass = openpyxl.Workbook()
    ws_pass = wb_pass.active
    ws_pass.title = "Passed Cases"
    for col, h in enumerate(headers1, 1):
        ws_pass.cell(1, col, h).font = Font(bold=True, color=WHITE)
        ws_pass.cell(1, col, h).fill = PatternFill("solid", fgColor=GREEN)
    
    r_idx = 2
    for t in report["tests"]:
        if t["outcome"] == "passed":
            ws_pass.cell(r_idx, 1, t["id"])
            ws_pass.cell(r_idx, 2, t["module"])
            ws_pass.cell(r_idx, 3, t["name"])
            ws_pass.cell(r_idx, 4, t["priority"])
            ws_pass.cell(r_idx, 5, "PASSED").font = Font(bold=True, color=GREEN)
            ws_pass.cell(r_idx, 6, t["duration"])
            r_idx += 1
    wb_pass.save(str(EXCEL_DIR / "Passed_Test_Cases.xlsx"))

    # -- Report 3: Failed_Test_Cases.xlsx ----------------------------------------
    wb_fail = openpyxl.Workbook()
    ws_fail = wb_fail.active
    ws_fail.title = "Failed Cases"
    for col, h in enumerate(headers_fail, 1):
        ws_fail.cell(1, col, h).font = Font(bold=True, color=WHITE)
        ws_fail.cell(1, col, h).fill = PatternFill("solid", fgColor=RED)
    
    r_idx = 2
    for t in report["tests"]:
        if t["outcome"] == "failed":
            ws_fail.cell(r_idx, 1, t["id"])
            ws_fail.cell(r_idx, 2, t["module"])
            ws_fail.cell(r_idx, 3, t["name"])
            ws_fail.cell(r_idx, 4, t["priority"])
            ws_fail.cell(r_idx, 5, "FAILED").font = Font(bold=True, color=RED)
            ws_fail.cell(r_idx, 6, t["duration"])
            ws_fail.cell(r_idx, 7, t["error"])
            r_idx += 1
    wb_fail.save(str(EXCEL_DIR / "Failed_Test_Cases.xlsx"))

    # -- Report 4: Execution_Summary.xlsx ----------------------------------------
    wb_sum = openpyxl.Workbook()
    ws_sum = wb_sum.active
    ws_sum.title = "Execution Summary"
    for col, h in enumerate(["Metric", "Value"], 1):
        ws_sum.cell(1, col, h).font = Font(bold=True, color=WHITE)
        ws_sum.cell(1, col, h).fill = PatternFill("solid", fgColor=DARK)
    for r, (m, v) in enumerate(metrics, 2):
        ws_sum.cell(r, 1, m).font = Font(bold=True)
        ws_sum.cell(r, 2, v)
    wb_sum.save(str(EXCEL_DIR / "Summary_Report.xlsx"))

# ─────────────────────────────────────────────────────────────
# HTML Dashboards Generation
# ─────────────────────────────────────────────────────────────
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1.0"/>
  <title>MediBook Appium Mobile Test Report — Build #{{ run_number }}</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0d0d1a;color:#e2e8f0;padding:2rem}
    h1{font-size:1.8rem;background:linear-gradient(135deg,#8b5cf6,#2563eb);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:.25rem}
    .subtitle{color:#64748b;font-size:.85rem;margin-bottom:2rem}
    .kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:1rem;margin-bottom:2rem}
    .kpi{background:#15152a;border-radius:12px;padding:1.25rem 1rem;text-align:center;border:1px solid #27274a;position:relative}
    .kpi .val{font-size:2rem;font-weight:800;line-height:1}
    .kpi.total .val{color:#a5b4fc}
    .kpi.pass .val{color:#10b981}
    .kpi.fail .val{color:#ef4444}
    .kpi.skip .val{color:#f59e0b}
    .kpi.rate .val{color:#8b5cf6}
    .kpi .unit{font-size:.75rem;color:#64748b;margin-top:.4rem;text-transform:uppercase;letter-spacing:.05em}
    .charts-grid{display:grid;grid-template-columns:1fr 1fr;gap:1.5rem;margin-bottom:2rem}
    @media(max-width:900px){.charts-grid{grid-template-columns:1fr}}
    .chart-card{background:#15152a;border-radius:12px;padding:1.5rem;border:1px solid #27274a}
    .chart-card h3{font-size:.85rem;color:#94a3b8;margin-bottom:1rem;text-transform:uppercase}
    table{width:100%;border-collapse:collapse;background:#15152a;border-radius:12px;overflow:hidden;margin-bottom:2rem;border:1px solid #27274a}
    th{background:#1a1a36;padding:.75rem 1rem;text-align:left;font-size:.7rem;text-transform:uppercase;color:#8888b0}
    td{padding:.75rem 1rem;border-bottom:1px solid #1f1f3a;font-size:.8rem}
    tr:last-child td{border-bottom:none}tr:hover td{background:#1d1d3b}
    .badge{display:inline-block;padding:.2rem .5rem;border-radius:9999px;font-size:.7rem;font-weight:700}
    .badge.passed{background:#044e3b;color:#10b981}
    .badge.failed{background:#450a0a;color:#ef4444}
    .badge.skipped{background:#452b0a;color:#f59e0b}
    .error-log{background:#2a0f10;border-left:3px solid #ef4444;color:#fca5a5;padding:.5rem;border-radius:4px;font-family:monospace;font-size:.7rem;margin-top:.5rem;white-space:pre-wrap;max-height:100px;overflow-y:auto}
  </style>
</head>
<body>
  <h1>MediBook Mobile Automation E2E — Dashboard</h1>
  <p class="subtitle">Appium Testing Portal · Android Viewport · Build #{{ run_number }} · {{ generated_at }}</p>

  <div class="kpi-grid">
    <div class="kpi total">
      <div class="val">{{ total }}</div>
      <div class="unit">Total cases</div>
    </div>
    <div class="kpi pass">
      <div class="val">{{ passed }}</div>
      <div class="unit">Passed</div>
    </div>
    <div class="kpi fail">
      <div class="val">{{ failed }}</div>
      <div class="unit">Failed</div>
    </div>
    <div class="kpi skip">
      <div class="val">{{ skipped }}</div>
      <div class="unit">Skipped</div>
    </div>
    <div class="kpi rate">
      <div class="val">{{ "%.1f"|format(passed/total*100) }}%</div>
      <div class="unit">Pass Rate</div>
    </div>
  </div>

  <div class="charts-grid">
    <div class="chart-card">
      <h3>Outcome Distribution</h3>
      <canvas id="outcomeChart" height="150"></canvas>
    </div>
    <div class="chart-card">
      <h3>Module Analysis</h3>
      <canvas id="moduleChart" height="150"></canvas>
    </div>
  </div>

  <h2>Execution Details</h2>
  <table>
    <thead>
      <tr>
        <th>Test ID</th><th>Module</th><th>Test Scenario Name</th><th>Priority</th><th>Status</th><th>Duration</th>
      </tr>
    </thead>
    <tbody>
      {% for t in tests %}
      <tr>
        <td><code style="color:#a5b4fc">{{ t.id }}</code></td>
        <td>{{ t.module }}</td>
        <td>
          <strong>{{ t.name }}</strong>
          {% if t.error %}
          <div class="error-log">{{ t.error }}</div>
          {% endif %}
        </td>
        <td>{{ t.priority }}</td>
        <td><span class="badge {{ t.outcome }}">{{ t.outcome.upper() }}</span></td>
        <td>{{ "%.2f"|format(t.duration) }}s</td>
      </tr>
      {% endfor %}
    </tbody>
  </table>

  <script>
    const ctx = document.getElementById('outcomeChart').getContext('2d');
    new Chart(ctx, {
      type: 'doughnut',
      data: {
        labels: ['Passed', 'Failed', 'Skipped'],
        datasets: [{
          data: [{{ passed }}, {{ failed }}, {{ skipped }}],
          backgroundColor: ['#10b981', '#ef4444', '#f59e0b'],
          borderColor: '#15152a',
          borderWidth: 2
        }]
      },
      options: {
        responsive: true,
        plugins: { legend: { labels: { color: '#94a3b8' } } }
      }
    });

    const mCtx = document.getElementById('moduleChart').getContext('2d');
    const modules = {{ modules | tojson }};
    const passedCounts = {{ passed_counts | tojson }};
    const failedCounts = {{ failed_counts | tojson }};

    new Chart(mCtx, {
      type: 'bar',
      data: {
        labels: modules,
        datasets: [
          { label: 'Passed', data: passedCounts, backgroundColor: '#10b981' },
          { label: 'Failed', data: failedCounts, backgroundColor: '#ef4444' }
        ]
      },
      options: {
        responsive: true,
        scales: {
          x: { stacked: true, ticks: { color: '#64748b', font: { size: 9 } } },
          y: { stacked: true, ticks: { color: '#64748b' } }
        },
        plugins: { legend: { labels: { color: '#94a3b8' } } }
      }
    });
  </script>
</body>
</html>"""

def generate_html_reports(report: dict, base_url: str, run_number: str):
    if not HAS_JINJA2:
        print("[WARN] jinja2 not installed - skipping HTML dashboards.")
        return

    # Extract module arrays for JavaScript charts
    module_stats = {}
    for t in report["tests"]:
        m = t["module"]
        if m not in module_stats:
            module_stats[m] = {"passed": 0, "failed": 0}
        if t["outcome"] == "passed":
            module_stats[m]["passed"] += 1
        elif t["outcome"] == "failed":
            module_stats[m]["failed"] += 1

    modules = list(module_stats.keys())
    passed_counts = [module_stats[m]["passed"] for m in modules]
    failed_counts = [module_stats[m]["failed"] for m in modules]

    env = Environment(loader=BaseLoader())
    tmpl = env.from_string(HTML_TEMPLATE)
    html = tmpl.render(
        **report,
        run_number=run_number,
        modules=modules,
        passed_counts=passed_counts,
        failed_counts=failed_counts
    )

    # 1. execution-report.html
    path1 = str(HTML_DIR / "execution-report.html")
    with open(path1, "w", encoding="utf-8") as f:
        f.write(html)
    
    # 2. dashboard.html
    path2 = str(HTML_DIR / "dashboard.html")
    with open(path2, "w", encoding="utf-8") as f:
        f.write(html)

    # 3. trends.html
    path3 = str(HTML_DIR / "trends.html")
    with open(path3, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"[OK] HTML dashboard -> {path1}")

# ─────────────────────────────────────────────────────────────
# Markdown Summary
# ─────────────────────────────────────────────────────────────
def generate_markdown_summary(report: dict, base_url: str, run_number: str):
    pct = round(report["passed"] / report["total"] * 100, 1) if report["total"] else 0
    fail_pct = round(100 - pct, 1)
    
    lines = [
        "# Android Appium E2E Execution Summary\n",
        f"**Build Number:** #{run_number}  ",
        f"**Execution Date:** {report['generated_at']}  ",
        "**Git Commit:** github.sha  ",
        "**Branch:** github.ref  \n",
        "**APK Version:** 1.0.0-debug  \n",
        "**Device:** emulator-5554  ",
        "**Android Version:** 13.0  \n",
        "## Execution Metrics\n",
        f"**Total Test Cases:** {report['total']}  ",
        f"**Executed:** {report['total']}  ",
        f"**Passed:** {report['passed']}  ",
        f"**Failed:** {report['failed']}  ",
        f"**Skipped:** {report['skipped']}  ",
        f"**Blocked:** {report['blocked']}  \n",
        f"**Pass Percentage:** **{pct}%**  ",
        f"**Fail Percentage:** **{fail_pct}%**  \n",
        f"**Execution Duration:** {report['duration']}s\n",
        "## VALID TEST CASE SUMMARY\n",
        "### PASSED TESTS\n"
    ]
    
    # Print a few passed test examples
    pass_examples = [t for t in report["tests"] if t["outcome"] == "passed"][:4]
    for t in pass_examples:
        lines.append(f"✓ `{t['id']}` — {t['name']}")
        
    lines.append("\n### FAILED TESTS\n")
    fail_examples = [t for t in report["tests"] if t["outcome"] == "failed"]
    for t in fail_examples:
        lines.append(f"✗ `{t['id']}` — {t['name']}")
        lines.append(f"  *Reason: {t['error'].strip()}*")

    lines.append("\n### SKIPPED TESTS\n")
    skip_examples = [t for t in report["tests"] if t["outcome"] == "skipped"]
    for t in skip_examples:
        lines.append(f"- `{t['id']}` — {t['name']}")
        lines.append(f"  *Reason: {t['error'].strip()}*")

    out_path = str(SUMMARY_DIR / "summary.md")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"[OK] Summary markdown -> {out_path}")

# ─────────────────────────────────────────────────────────────
# Main entry
# ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--json",       default="test-results.json")
    parser.add_argument("--base-url",   default="http://localhost:5173")
    parser.add_argument("--run-number", default="local")
    args = parser.parse_args()

    if os.path.exists(args.json):
        report = parse_results(args.json)
        generate_excel_reports(report, args.base_url, args.run_number)
        generate_html_reports(report,  args.base_url, args.run_number)
        generate_markdown_summary(report, args.base_url, args.run_number)
        print(f"\n[DONE] All reports exported successfully in separate folders under: {RESULTS_DIR}")
    else:
        print(f"[ERROR] Pytest JSON result file not found at: {args.json}")

if __name__ == "__main__":
    main()
